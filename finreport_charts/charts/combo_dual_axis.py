from __future__ import annotations

from pathlib import Path

import pandas as pd

from ..utils.mpl_style import apply_pretty_style
from ..utils.numfmt import choose_unit_scale, fmt_scaled, fmt_tick


def render_combo_png(
    df: pd.DataFrame,
    *,
    title: str,
    x_col: str,
    bar_col: str,
    line_col: str,
    out_png: Path,
    bar_label: str = "",
    line_label: str = "",
    x_label: str = "",
):
    import matplotlib.pyplot as plt

    apply_pretty_style()

    x = df[x_col].astype(str).tolist()
    y1 = df[bar_col].tolist()
    y2 = df[line_col].tolist()

    # unit for bar axis
    max_abs = 0.0
    try:
        max_abs = float(pd.Series(y1).abs().max())
    except Exception:
        max_abs = 0.0
    us = choose_unit_scale(max_abs)

    fig, ax1 = plt.subplots(figsize=(max(7, min(22, 0.75 * len(x) + 4)), 5.5))
    cont = ax1.bar(x, y1, color="#4E79A7", alpha=0.9, label=bar_label or bar_col)
    ax1.set_xlabel(x_label or "时间")
    ax1.set_ylabel(bar_label or bar_col)
    ax1.tick_params(axis="x", rotation=45)

    from matplotlib.ticker import FuncFormatter

    ax1.yaxis.set_major_formatter(FuncFormatter(lambda v, _pos: fmt_tick(v, us)))

    # value labels on bars
    ymax = max([float(v) for v in y1 if v is not None] + [0.0])
    ymin = min([float(v) for v in y1 if v is not None] + [0.0])
    span = max(1.0, ymax - ymin)
    ax1.set_ylim(ymin - 0.08 * span, ymax + 0.18 * span)

    for p in cont.patches:
        h = p.get_height()
        x0 = p.get_x() + p.get_width() / 2
        txt = fmt_scaled(float(h), us)
        if h >= 0:
            ax1.text(x0, h, txt, ha="center", va="bottom", fontsize=9, color="#EAEAEA")
        else:
            ax1.text(x0, h, txt, ha="center", va="top", fontsize=9, color="#EAEAEA")

    ax2 = ax1.twinx()
    ax2.plot(x, y2, color="#FF7F0E", marker="o", linewidth=2, label=line_label or line_col)
    ax2.set_ylabel(line_label or line_col)

    ax1.set_title(title)

    # 合并图例
    h1, l1 = ax1.get_legend_handles_labels()
    h2, l2 = ax2.get_legend_handles_labels()
    ax1.legend(h1 + h2, l1 + l2, loc="upper left")

    fig.tight_layout()
    out_png.parent.mkdir(parents=True, exist_ok=True)
    fig.savefig(out_png, dpi=200)
    plt.close(fig)


def write_combo_excel(
    df: pd.DataFrame,
    *,
    title: str,
    x_col: str,
    bar_col: str,
    line_col: str,
    out_xlsx: Path,
    bar_label: str = "",
    line_label: str = "",
    x_label: str = "",
):
    from openpyxl import Workbook
    from openpyxl.chart import BarChart, LineChart, Reference
    from openpyxl.chart.axis import ChartLines
    from openpyxl.styles import Alignment, Font, PatternFill

    from ..utils.xlsx import autofit_columns

    out_xlsx.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "data"

    ws["A1"].value = title
    ws.merge_cells("A1:C1")
    ws["A1"].font = Font(bold=True, size=14, color="FFFFFF")
    ws["A1"].fill = PatternFill("solid", fgColor="0B2F4F")
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 24

    ws.append([x_col, bar_col, line_col])
    for _, r in df.iterrows():
        ws.append([str(r[x_col]), float(r[bar_col]) if pd.notna(r[bar_col]) else None, float(r[line_col]) if pd.notna(r[line_col]) else None])

    # number formats
    for row in ws.iter_rows(min_row=3, min_col=2, max_col=3, max_row=ws.max_row):
        for cell in row:
            if isinstance(cell.value, (int, float)):
                cell.number_format = "#,##0.00" if cell.column == 3 else "#,##0"

    autofit_columns(ws)

    ws_chart = wb.create_sheet("chart")
    ws_chart["A1"].value = title
    ws_chart["A1"].font = Font(bold=True, size=14)

    # Bar chart
    bar = BarChart()
    bar.type = "col"
    bar.title = title
    bar.y_axis.title = bar_label or bar_col
    bar.x_axis.title = x_label or "时间"

    bar_data = Reference(ws, min_col=2, min_row=2, max_row=ws.max_row)
    cats = Reference(ws, min_col=1, min_row=3, max_row=ws.max_row)
    bar.add_data(bar_data, titles_from_data=True)
    bar.set_categories(cats)

    # Line chart on secondary axis
    line = LineChart()
    line.y_axis.axId = 200
    line.y_axis.title = line_label or line_col
    line.y_axis.crosses = "max"
    line_data = Reference(ws, min_col=3, min_row=2, max_row=ws.max_row)
    line.add_data(line_data, titles_from_data=True)

    bar += line
    bar.width = 24
    bar.height = 12

    ws_chart.add_chart(bar, "A3")

    wb.save(out_xlsx)

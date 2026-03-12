from __future__ import annotations

from pathlib import Path

import pandas as pd

from ..utils.mpl_style import apply_pretty_style


def topn_with_other(items: list[tuple[str, float]], top_n: int) -> list[tuple[str, float]]:
    items2 = [(k, float(v)) for k, v in items if v is not None]
    items2.sort(key=lambda x: abs(x[1]), reverse=True)
    head = items2[:top_n]
    tail = items2[top_n:]
    if not tail:
        return head
    other = sum(v for _, v in tail)
    head.append(("其他", other))
    return head


def render_pie_png(items: list[tuple[str, float]], *, title: str, out_png: Path):
    import matplotlib.pyplot as plt

    apply_pretty_style()

    labels = [k for k, _ in items]
    values = [v for _, v in items]

    fig, ax = plt.subplots()
    ax.pie(values, labels=labels, autopct=lambda p: f"{p:.1f}%" if p >= 3 else "")
    ax.set_title(title)
    fig.tight_layout()

    out_png.parent.mkdir(parents=True, exist_ok=True)
    fig.savefig(out_png, dpi=200)
    plt.close(fig)


def write_pie_excel(items: list[tuple[str, float]], *, title: str, out_xlsx: Path):
    from openpyxl import Workbook
    from openpyxl.chart import PieChart, Reference
    from openpyxl.styles import Alignment, Font, PatternFill

    out_xlsx.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "data"

    ws["A1"].value = title
    ws.merge_cells("A1:B1")
    ws["A1"].font = Font(bold=True, size=14, color="FFFFFF")
    ws["A1"].fill = PatternFill("solid", fgColor="0B2F4F")
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 24

    ws.append(["科目", "数值"])
    for k, v in items:
        ws.append([k, float(v)])

    for row in ws.iter_rows(min_row=3, min_col=2, max_col=2, max_row=ws.max_row):
        for cell in row:
            if isinstance(cell.value, (int, float)):
                cell.number_format = "#,##0"

    ws_chart = wb.create_sheet("chart")
    ws_chart["A1"].value = title
    ws_chart["A1"].font = Font(bold=True, size=14)

    chart = PieChart()
    chart.title = title
    data = Reference(ws, min_col=2, min_row=2, max_row=ws.max_row)
    labels = Reference(ws, min_col=1, min_row=3, max_row=ws.max_row)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(labels)
    chart.width = 18
    chart.height = 12

    ws_chart.add_chart(chart, "A3")

    wb.save(out_xlsx)

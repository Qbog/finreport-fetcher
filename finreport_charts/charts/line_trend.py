from __future__ import annotations

from pathlib import Path
import math

import pandas as pd

from ..utils.mpl_style import apply_pretty_style
from ..utils.numfmt import UnitScale, choose_unit_scale, fmt_tick


def render_lines_png(
    df: pd.DataFrame,
    *,
    title: str,
    x_col: str,
    series: list[tuple[str, str]],
    out_png: Path,
    x_label: str = "时间",
    y_label: str = "",
    y_range: tuple[float, float] | None = None,
    unit_scale: UnitScale | None = None,
    figsize: tuple[float, float] | None = None,
    mark_dates: list[str] | None = None,
    max_xticks: int = 12,
):
    """Multi-series line chart.

    series: [(col_name_in_df, display_label), ...]
    """

    import matplotlib.pyplot as plt
    import matplotlib.dates as mdates
    from matplotlib.ticker import FuncFormatter

    apply_pretty_style()

    # x axis: try parse as datetime for smart tick formatting
    x_raw = df[x_col]
    x_dt = pd.to_datetime(x_raw, errors="coerce")
    use_dt = bool(x_dt.notna().sum() >= 2)
    x = x_dt.to_numpy() if use_dt else x_raw.astype(str).tolist()
    n = len(x)
    k = max(1, len(series))

    # compute unit
    max_abs = 0.0
    for col, _label in series:
        try:
            v = float(pd.to_numeric(df[col], errors="coerce").abs().max())
            if math.isfinite(v):
                max_abs = max(max_abs, v)
        except Exception:
            pass
    us = unit_scale or choose_unit_scale(max_abs)

    fig_size = figsize or (max(7, min(22, 0.75 * n + 4)), 5.5)
    fig, ax = plt.subplots(figsize=fig_size)

    palette = [
        "#4E79A7",
        "#F28E2B",
        "#E15759",
        "#76B7B2",
        "#59A14F",
        "#EDC949",
        "#AF7AA1",
        "#FF9DA7",
        "#9C755F",
        "#BAB0AC",
    ]

    for j, (col, label) in enumerate(series):
        y = pd.to_numeric(df[col], errors="coerce").tolist()
        ax.plot(
            x,
            y,
            color=palette[j % len(palette)],
            marker=None,
            linewidth=2,
            label=label or col,
        )

    ax.set_title(title)
    ax.set_xlabel(x_label or "时间")
    ax.set_ylabel(y_label)

    if use_dt:
        locator = mdates.AutoDateLocator(minticks=6, maxticks=max(6, int(max_xticks or 12)))
        ax.xaxis.set_major_locator(locator)
        # show month explicitly; concise formatter will include month/day when appropriate
        ax.xaxis.set_major_formatter(mdates.ConciseDateFormatter(locator))
        fig.autofmt_xdate(rotation=30)

        # optional quarter-end markers or other key dates
        for ds in mark_dates or []:
            try:
                d0 = pd.to_datetime(ds, errors="coerce")
                if pd.isna(d0):
                    continue
                ax.axvline(d0, color="#999999", linestyle="--", linewidth=1, alpha=0.35)
            except Exception:
                continue
    else:
        # category x axis: rotate slightly but avoid printing every label manually
        ax.tick_params(axis="x", rotation=30)
    ax.yaxis.set_major_formatter(FuncFormatter(lambda v, _pos: fmt_tick(v, us)))
    if y_range and len(y_range) == 2:
        ax.set_ylim(y_range[0], y_range[1])

    if k > 1:
        ax.legend(loc="best")

    fig.tight_layout()
    out_png.parent.mkdir(parents=True, exist_ok=True)
    fig.savefig(out_png, dpi=220)
    plt.close(fig)


def write_lines_excel(
    df: pd.DataFrame,
    *,
    title: str,
    x_col: str,
    series: list[tuple[str, str]],
    out_xlsx: Path,
    x_label: str = "时间",
    y_label: str = "",
    y_range: tuple[float, float] | None = None,
):
    """Multi-series line chart exported to Excel.

    series: [(col_name_in_df, display_label), ...]
    """

    from openpyxl import Workbook
    from openpyxl.chart import LineChart, Reference
    from openpyxl.styles import Alignment, Font, PatternFill

    from ..utils.xlsx import autofit_columns

    out_xlsx.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "data"

    # 标题
    ncols = 1 + max(1, len(series))
    ws["A1"].value = title
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
    ws["A1"].font = Font(bold=True, size=14, color="FFFFFF")
    ws["A1"].fill = PatternFill("solid", fgColor="0B2F4F")
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 24

    # 表头
    header = [x_col] + [label for _, label in series]
    ws.append(header)

    # 数据
    for _, r in df.iterrows():
        row = [str(r[x_col])]
        for col, _label in series:
            v = r.get(col)
            row.append(float(v) if pd.notna(v) else None)
        ws.append(row)

    # number formats
    for row in ws.iter_rows(min_row=3, min_col=2, max_col=ncols, max_row=ws.max_row):
        for cell in row:
            if isinstance(cell.value, (int, float)):
                cell.number_format = "#,##0"

    autofit_columns(ws)

    # 图表
    ws_chart = wb.create_sheet("chart")
    ws_chart["A1"].value = title
    ws_chart["A1"].font = Font(bold=True, size=14)

    chart = LineChart()
    chart.title = title
    chart.y_axis.title = y_label
    chart.x_axis.title = x_label or "时间"
    chart.marker = None  # avoid markers (too crowded)

    data = Reference(ws, min_col=2, min_row=2, max_col=ncols, max_row=ws.max_row)
    cats = Reference(ws, min_col=1, min_row=3, max_row=ws.max_row)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.width = 22
    chart.height = 12

    if y_range and len(y_range) == 2:
        chart.y_axis.scaling.min = y_range[0]
        chart.y_axis.scaling.max = y_range[1]

    ws_chart.add_chart(chart, "A3")

    wb.save(out_xlsx)

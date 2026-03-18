from __future__ import annotations

from pathlib import Path
import math

import pandas as pd

from ..utils.mpl_style import apply_pretty_style
from ..utils.numfmt import UnitScale, choose_unit_scale, fmt_scaled, fmt_tick


def render_bar_png(
    df: pd.DataFrame,
    *,
    title: str,
    x_col: str,
    y_col: str,
    out_png: Path,
    y_label: str = "",
    x_label: str = "",
    y_range: tuple[float, float] | None = None,
    unit_scale: UnitScale | None = None,
    figsize: tuple[float, float] | None = None,
):
    """Single-series bar chart (backward compatible)."""

    render_bars_png(
        df,
        title=title,
        x_col=x_col,
        series=[(y_col, y_col)],
        out_png=out_png,
        x_label=x_label,
        y_label=y_label or y_col,
        y_range=y_range,
        unit_scale=unit_scale,
        figsize=figsize,
    )


def render_bars_png(
    df: pd.DataFrame,
    *,
    title: str,
    x_col: str,
    series: list[tuple[str, str]],
    out_png: Path,
    x_label: str = "时间",
    y_label: str = "",
    series_colors: list[str | None] | None = None,
    x_colors: list[str | None] | None = None,
    y_range: tuple[float, float] | None = None,
    unit_scale: UnitScale | None = None,
    figsize: tuple[float, float] | None = None,
):
    """Multi-series (grouped) bar chart.

    series: [(col_name_in_df, display_label), ...]
    """

    import matplotlib.pyplot as plt
    import numpy as np
    from matplotlib.ticker import FuncFormatter

    apply_pretty_style()

    x = df[x_col].astype(str).tolist()
    n = len(x)
    k = max(1, len(series))

    idx = np.arange(n)
    width = 0.8 / k

    # compute unit
    max_abs = 0.0
    for col, _label in series:
        try:
            v = float(pd.to_numeric(df[col], errors="coerce").abs().max())
            if math.isfinite(v):
                max_abs = max(max_abs, v)
        except Exception:
            pass
    us = unit_scale or choose_unit_scale(max_abs)

    fig_size = figsize or (max(7, min(22, 0.75 * n + 4)), 5.5)
    fig, ax = plt.subplots(figsize=fig_size)

    palette = [
        "#4E79A7",
        "#F28E2B",
        "#E15759",
        "#76B7B2",
        "#59A14F",
        "#EDC949",
        "#AF7AA1",
        "#FF9DA7",
        "#9C755F",
        "#BAB0AC",
    ]

    def _norm_hex(c: str | None) -> str | None:
        if not c:
            return None
        s = str(c).strip()
        if not s:
            return None
        if s.startswith("#"):
            s = s[1:]
        if len(s) == 6:
            return f"#{s}"
        return None

    containers = []
    for j, (col, label) in enumerate(series):
        y = pd.to_numeric(df[col], errors="coerce").tolist()
        offset = (j - (k - 1) / 2) * width

        # color policy:
        # - multi-series: per-series color (series_colors[j]) or palette
        # - single-series: allow per-x color list (x_colors)
        if k == 1 and x_colors:
            colors = [_norm_hex(c) or palette[0] for c in x_colors]
            cont = ax.bar(
                idx + offset,
                y,
                width=width,
                label=label,
                color=colors,
                alpha=0.92,
            )
        else:
            c0 = None
            if series_colors and j < len(series_colors):
                c0 = _norm_hex(series_colors[j])
            cont = ax.bar(
                idx + offset,
                y,
                width=width,
                label=label,
                color=c0 or palette[j % len(palette)],
                alpha=0.92,
            )
        containers.append(cont)

    ax.set_title(title)
    ax.set_xlabel(x_label or "时间")
    ax.set_ylabel(y_label)
    ax.set_xticks(idx)
    ax.set_xticklabels(x, rotation=45, ha="right")

    # y-axis ticks with unit
    ax.yaxis.set_major_formatter(FuncFormatter(lambda v, _pos: fmt_tick(v, us)))

    # add value labels above bars
    ymax = None
    ymin = None
    for cont in containers:
        for p in cont.patches:
            h = p.get_height()
            if h is None or (isinstance(h, float) and not math.isfinite(h)):
                continue
            ymax = float(h) if ymax is None else max(ymax, float(h))
            ymin = float(h) if ymin is None else min(ymin, float(h))

    # headroom (override when y_range is provided)
    if y_range and len(y_range) == 2:
        ax.set_ylim(y_range[0], y_range[1])
    else:
        ymax0 = ymax if ymax is not None else 0.0
        ymin0 = ymin if ymin is not None else 0.0
        span = max(1.0, ymax0 - ymin0)
        ax.set_ylim(ymin0 - 0.08 * span, ymax0 + 0.18 * span)

    for cont in containers:
        for p in cont.patches:
            h = p.get_height()
            if h is None or (isinstance(h, float) and not math.isfinite(h)):
                continue
            x0 = p.get_x() + p.get_width() / 2
            txt = fmt_scaled(float(h), us)
            if h >= 0:
                ax.text(x0, h, txt, ha="center", va="bottom", fontsize=9, color="#EAEAEA")
            else:
                ax.text(x0, h, txt, ha="center", va="top", fontsize=9, color="#EAEAEA")

    if k > 1:
        ax.legend(loc="best")

    fig.tight_layout()
    out_png.parent.mkdir(parents=True, exist_ok=True)
    fig.savefig(out_png, dpi=220)
    plt.close(fig)


def write_bar_excel(
    df: pd.DataFrame,
    *,
    title: str,
    x_col: str,
    y_col: str,
    out_xlsx: Path,
    y_label: str = "",
    x_label: str = "",
    y_range: tuple[float, float] | None = None,
):
    """Single-series Excel output (backward compatible)."""

    write_bars_excel(
        df,
        title=title,
        x_col=x_col,
        series=[(y_col, y_col)],
        out_xlsx=out_xlsx,
        x_label=x_label,
        y_label=y_label or y_col,
        y_range=y_range,
    )


def write_bars_excel(
    df: pd.DataFrame,
    *,
    title: str,
    x_col: str,
    series: list[tuple[str, str]],
    out_xlsx: Path,
    x_label: str = "时间",
    y_label: str = "",
    series_colors: list[str | None] | None = None,
    x_colors: list[str | None] | None = None,
    y_range: tuple[float, float] | None = None,
):
    """Multi-series bar chart exported to Excel.

    series: [(col_name_in_df, display_label), ...]
    """

    from openpyxl import Workbook
    from openpyxl.chart import BarChart, Reference
    from openpyxl.styles import Alignment, Font, PatternFill

    from ..utils.xlsx import autofit_columns

    out_xlsx.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "data"

    # 标题
    # 写入宽度取决于列数
    ncols = 1 + max(1, len(series))
    ws["A1"].value = title
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
    ws["A1"].font = Font(bold=True, size=14, color="FFFFFF")
    ws["A1"].fill = PatternFill("solid", fgColor="0B2F4F")
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 24

    # 表头
    header = [x_col] + [label for _, label in series]
    ws.append(header)

    # 数据
    for _, r in df.iterrows():
        row = [str(r[x_col])]
        for col, _label in series:
            v = r.get(col)
            row.append(float(v) if pd.notna(v) else None)
        ws.append(row)

    # number formats
    for row in ws.iter_rows(min_row=3, min_col=2, max_col=ncols, max_row=ws.max_row):
        for cell in row:
            if isinstance(cell.value, (int, float)):
                cell.number_format = "#,##0"

    autofit_columns(ws)

    # 图表
    ws_chart = wb.create_sheet("chart")
    ws_chart["A1"].value = title
    ws_chart["A1"].font = Font(bold=True, size=14)

    chart = BarChart()
    chart.type = "col"
    chart.title = title
    chart.y_axis.title = y_label
    chart.x_axis.title = x_label or "时间"

    data = Reference(ws, min_col=2, min_row=2, max_col=ncols, max_row=ws.max_row)
    cats = Reference(ws, min_col=1, min_row=3, max_row=ws.max_row)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.width = 22
    chart.height = 12

    if y_range and len(y_range) == 2:
        chart.y_axis.scaling.min = y_range[0]
        chart.y_axis.scaling.max = y_range[1]

    def _hex6(c: str | None) -> str | None:
        if not c:
            return None
        s = str(c).strip()
        if not s:
            return None
        if s.startswith("#"):
            s = s[1:]
        if len(s) == 6:
            return s.upper()
        return None

    # Apply colors
    try:
        from openpyxl.chart.series import DataPoint

        for j, s in enumerate(chart.series):
            c0 = _hex6(series_colors[j]) if series_colors and j < len(series_colors) else None
            if c0:
                s.graphicalProperties.solidFill = c0

            # per-point colors for single-series compare charts
            if len(chart.series) == 1 and x_colors:
                s.dPt = []
                for i, xc in enumerate(x_colors):
                    cc = _hex6(xc)
                    if not cc:
                        continue
                    dp = DataPoint(idx=i)
                    dp.graphicalProperties.solidFill = cc
                    s.dPt.append(dp)
    except Exception:
        pass

    ws_chart.add_chart(chart, "A3")

    wb.save(out_xlsx)

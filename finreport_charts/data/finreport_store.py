from __future__ import annotations

import re
import subprocess
import sys
from dataclasses import dataclass
from datetime import date
from pathlib import Path

import pandas as pd

from finreport_fetcher.utils.dates import parse_date, quarter_ends_between
from finreport_fetcher.utils.paths import safe_dir_component


@dataclass(frozen=True)
class FinreportPaths:
    xlsx: Path
    pdf: Path | None


def _company_root_dir(data_dir: Path, code6: str, name: str | None = None) -> Path:
    """Resolve company root directory under data_dir."""

    if name:
        cand = data_dir / safe_dir_component(f"{name}_{code6}")
        if cand.exists() and cand.is_dir():
            return cand

    matches = sorted([p for p in data_dir.glob(f"*_{code6}") if p.is_dir()])
    if matches:
        return matches[0]

    return data_dir


def _company_reports_dir(data_dir: Path, code6: str, name: str | None = None) -> Path:
    """Resolve reports directory under data_dir.

    Layout:
      data_dir/{公司名}_{code6}/reports/*.xlsx

    Backward compatible:
      data_dir/{公司名}_{code6}/*.xlsx
      data_dir/*.xlsx (legacy)
    """

    root = _company_root_dir(data_dir, code6, name=name)
    cand = root / "reports"
    if cand.exists() and cand.is_dir():
        return cand
    return root


def expected_xlsx_path(data_dir: Path, code6: str, statement_type: str, period_end: date, *, name: str | None = None) -> Path:
    fname = f"{code6}_{statement_type}_{period_end.strftime('%Y%m%d')}.xlsx"

    # legacy flat layout
    p0 = data_dir / fname
    if p0.exists():
        return p0

    return _company_reports_dir(data_dir, code6, name=name) / fname


def expected_pdf_path(data_dir: Path, code6: str, period_end: date, *, name: str | None = None) -> Path:
    fname = f"{code6}_{period_end.strftime('%Y%m%d')}.pdf"

    # legacy
    p0 = data_dir / "pdf" / fname
    if p0.exists():
        return p0

    # new layout: company_dir/pdf/*.pdf
    return _company_root_dir(data_dir, code6, name=name) / "pdf" / fname


def ensure_finreports(
    *,
    code_or_name_args: list[str],
    code6: str,
    start: date,
    end: date,
    data_dir: Path,
    provider: str,
    statement_type: str,
    pdf: bool,
    company_name: str | None = None,
    tushare_token: str | None = None,
) -> list[date]:
    """确保 data_dir 里存在 start~end 的所有报告期末日财报 xlsx。

    返回：本次“仍然缺失”的报告期末日列表（如果补齐失败）。

    设计要点：
    - 只尝试下载“缺失的报告期”，避免重复下载/覆盖已有文件。
    - 即使补齐失败（例如最新一期尚未披露、或本地文件被占用导致 Permission denied），
      也不在这里抛异常；由上层决定是否 strict。
    """

    periods = quarter_ends_between(start, end)
    missing = [
        pe
        for pe in periods
        if not expected_xlsx_path(data_dir, code6, statement_type, pe, name=company_name).exists()
    ]

    if not missing:
        return []

    # 逐期补齐：避免 fetcher 对整个区间重复抓取/重复写文件。
    for pe in missing:
        args = [
            sys.executable,
            "-m",
            "finreport_fetcher",
            "fetch",
            *code_or_name_args,
            "--start",
            pe.strftime("%Y-%m-%d"),
            "--end",
            pe.strftime("%Y-%m-%d"),
            "--provider",
            provider,
            "--statement-type",
            statement_type,
            "--out",
            str(data_dir),
            "--no-clean",
        ]
        if pdf:
            args.append("--pdf")
        if tushare_token:
            args += ["--tushare-token", tushare_token]

        # 不用 check_call：单期失败时继续其他期，并把缺失留给 still_missing。
        # 同时捕获输出，避免在 finreport_charts 场景下打印冗长 traceback。
        res = subprocess.run(
            args,
            check=False,
            stdout=subprocess.PIPE,
            stderr=subprocess.PIPE,
            text=True,
        )
        if res.returncode != 0:
            # 只打印最后一行错误，既能定位问题，又不会刷屏。
            err = (res.stderr or "").strip().splitlines()
            tail = err[-1] if err else "(no stderr)"
            print(
                f"提示：补数失败 {code6} {pe.strftime('%Y-%m-%d')} (rc={res.returncode}): {tail}",
                file=sys.stderr,
            )

    still_missing = [
        pe
        for pe in periods
        if not expected_xlsx_path(data_dir, code6, statement_type, pe, name=company_name).exists()
    ]
    return still_missing


_PROVIDER_RE = re.compile(r"数据源:\s*([0-9A-Za-z_]+)")


def read_sheet_provider(xlsx_path: Path, sheet_name: str) -> str | None:
    """Read provider tag from sheet title row (cell A1).

    Expected format contains: '数据源: <provider>'
    """

    try:
        from openpyxl import load_workbook

        wb = load_workbook(xlsx_path, read_only=True, data_only=True)
        ws = wb[sheet_name] if sheet_name in wb.sheetnames else wb.active
        v = ws["A1"].value
        wb.close()
        s = str(v or "")
        m = _PROVIDER_RE.search(s)
        return m.group(1) if m else None
    except Exception:
        return None


def read_statement_df(xlsx_path: Path, sheet_name: str) -> pd.DataFrame:
    """读取某一期财报的某张表。

    兼容两种格式：
    - 老格式：列为 [科目, 数值]
    - 新格式：列包含 key / 科目(中英展示) / 数值

    返回 df 至少包含列: 科目, 数值；若存在也会保留 key。
    """

    # 我们的 xlsx：第1行标题、第2行注释、第3行表头
    df = pd.read_excel(xlsx_path, sheet_name=sheet_name, header=2)

    subj_raw = "科目_CN" if "科目_CN" in df.columns else "科目"

    keep: list[str] = []
    if "key" in df.columns:
        keep.append("key")
    keep += [subj_raw, "数值"]

    df2 = df[keep].copy()
    if subj_raw != "科目":
        df2.rename(columns={subj_raw: "科目"}, inplace=True)

    return df2


def _subject_cn_from_display(s: str) -> str:
    """Extract CN subject from display like '中文 (English)'."""

    ss = (s or "").strip()
    if " (" in ss and ss.endswith(")"):
        return ss.split(" (", 1)[0].strip()
    return ss


def get_item_value(xlsx_path: Path, sheet_name: str, item: str) -> float | None:
    """按科目取值。

    item 支持：
    - CN 科目名（精确匹配）
    - 模板 key（如 is.revenue / bs.cash），当 xlsx 含 key 列时生效

    兼容：xlsx 的“科目”列可能为 '中文 (English)'。

    重要：当 xlsx 不存在时返回 None（上层可选择跳过该报告期）。
    """

    if not xlsx_path.exists():
        return None

    df = read_statement_df(xlsx_path, sheet_name)

    item_s = str(item)
    if "key" in df.columns and "." in item_s:
        m = df["key"].astype(str) == item_s
    else:
        subj = df["科目"].astype(str).map(_subject_cn_from_display)
        m = subj == item_s

    sub = df[m]
    if sub.empty:
        return None
    v = sub.iloc[0]["数值"]
    try:
        if pd.isna(v):
            return None
        return float(v)
    except Exception:
        return None


def get_section_items(
    xlsx_path: Path, sheet_name: str, section: str
) -> list[tuple[str, float]]:
    """按 section（标题科目）取其后连续的子项，直到下一个标题行。

    标题行的识别规则：数值为空且科目非空。

    当 xlsx 不存在时返回空列表。
    """

    if not xlsx_path.exists():
        return []

    df = read_statement_df(xlsx_path, sheet_name)
    subj = df["科目"].astype(str)
    val = df["数值"]

    subj_cn = subj.map(_subject_cn_from_display)

    idxs = df.index[subj_cn == section].tolist()
    if not idxs:
        return []

    start_i = idxs[0] + 1
    items: list[tuple[str, float]] = []

    for i in range(start_i, len(df)):
        name = str(subj.iloc[i])
        name_cn = _subject_cn_from_display(name)
        v = val.iloc[i]

        # 下一标题
        if name_cn and (pd.isna(v) or v is None):
            break

        if not name:
            continue
        if pd.isna(v) or v is None:
            continue
        try:
            items.append((name, float(v)))
        except Exception:
            continue

    return items


def load_price_csv(price_csv: Path) -> pd.DataFrame:
    df = pd.read_csv(price_csv)
    # 约定列名 date, close
    if "date" not in df.columns or "close" not in df.columns:
        raise ValueError("股价 CSV 需要包含列: date, close")
    df["date"] = pd.to_datetime(df["date"]).dt.date
    df["close"] = pd.to_numeric(df["close"], errors="coerce")
    return df.dropna(subset=["close"]).sort_values("date")


def price_on_or_before(df_price: pd.DataFrame, dt: date) -> float | None:
    sub = df_price[df_price["date"] <= dt]
    if sub.empty:
        return None
    return float(sub.iloc[-1]["close"])

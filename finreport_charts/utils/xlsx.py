from __future__ import annotations

from openpyxl.utils import get_column_letter

import unicodedata


def _display_len(s: str) -> int:
    """Estimate display width (treat East Asian wide chars as 2)."""

    total = 0
    for ch in s:
        if unicodedata.east_asian_width(ch) in {"W", "F"}:
            total += 2
        else:
            total += 1
    return total


def autofit_columns(ws, *, min_width: int = 8, max_width: int = 60, padding: int = 2):
    """Rudimentary column auto-fit for openpyxl worksheets."""

    widths: dict[int, int] = {}
    for row in ws.iter_rows(values_only=True):
        for j, v in enumerate(row, start=1):
            if v is None:
                continue
            s = str(v)
            # cap very long strings
            if len(s) > 200:
                s = s[:200]
            widths[j] = max(widths.get(j, 0), _display_len(s))

    for j, w in widths.items():
        w2 = max(min_width, min(max_width, w + padding))
        ws.column_dimensions[get_column_letter(j)].width = w2

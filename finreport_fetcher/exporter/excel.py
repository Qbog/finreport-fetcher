from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path

import pandas as pd


@dataclass(frozen=True)
class ExcelExportResult:
    path: str


def _autofit_worksheet(ws):
    """简单自适应列宽：按单元格字符串长度估算。

    注意：存在合并单元格时，ws.columns 可能返回 MergedCell（没有 column_letter）。
    这里改为按列索引遍历，跳过 MergedCell。
    """

    from openpyxl.cell.cell import MergedCell
    from openpyxl.utils import get_column_letter

    for col_idx in range(1, ws.max_column + 1):
        max_len = 0
        for row_idx in range(1, ws.max_row + 1):
            cell = ws.cell(row_idx, col_idx)
            if isinstance(cell, MergedCell):
                continue
            v = cell.value
            if v is None:
                continue
            s = str(v)
            if len(s) > max_len:
                max_len = len(s)

        col_letter = get_column_letter(col_idx)
        # True-ish autofit: small minimum, larger maximum
        ws.column_dimensions[col_letter].width = min(max(4, max_len + 2), 80)


def export_bundle_to_excel(
    out_path: Path,
    balance_sheet: pd.DataFrame,
    income_statement: pd.DataFrame,
    cashflow_statement: pd.DataFrame,
    meta: dict,
    title_info: dict | None = None,
):
    """导出为 Excel（每张表一个 sheet），并进行美化。

    规则：
    - 不再输出“报告期末日”列；改为每个 sheet 顶部标题行展示。
    - 不再输出 PDF 链接/本地路径列；改为标题/注释展示（来源一致）。
    - 支持科目分组缩进：读取 df 中可选列 __level / __is_header。
    """

    out_path.parent.mkdir(parents=True, exist_ok=True)

    from ..mappings.enrich import enrich_statement_df

    bs_df = enrich_statement_df(balance_sheet, sheet_name_cn="资产负债表")
    is_df = enrich_statement_df(income_statement, sheet_name_cn="利润表")
    cf_df = enrich_statement_df(cashflow_statement, sheet_name_cn="现金流量表")

    def view_df(df: pd.DataFrame) -> pd.DataFrame:
        # 兼容老格式：至少输出 key/科目/数值（用户要求不输出 科目_CN/科目_EN）
        preferred = ["key", "科目", "数值"]
        cols = [c for c in preferred if c in df.columns]
        if not cols:
            cols = [c for c in ["科目", "数值"] if c in df.columns]

        out = df[cols].copy()

        # 在“数值”右侧插入空白列，方便肉眼查看数字列
        if "数值" in out.columns:
            idx = out.columns.get_loc("数值") + 1
            spacer_name = " "  # 单空格列名（Excel 表头看起来像空列）
            if spacer_name not in out.columns:
                out.insert(idx, spacer_name, [""] * len(out))

        return out

    # 写入：预留两行做标题/注释
    with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
        view_df(bs_df).to_excel(writer, sheet_name="资产负债表", index=False, startrow=2)
        view_df(is_df).to_excel(writer, sheet_name="利润表", index=False, startrow=2)
        view_df(cf_df).to_excel(writer, sheet_name="现金流量表", index=False, startrow=2)

    # 美化（openpyxl）
    from openpyxl import load_workbook
    from openpyxl.formatting.rule import CellIsRule
    from openpyxl.styles import Alignment, Font, PatternFill

    wb = load_workbook(out_path)

    title_fill = PatternFill("solid", fgColor="0B2F4F")
    title_font = Font(color="FFFFFF", bold=True, size=14)
    title_alignment = Alignment(horizontal="left", vertical="center")

    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(color="FFFFFF", bold=True)
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    zebra_fill = PatternFill("solid", fgColor="F7F7F7")
    section_fill = PatternFill("solid", fgColor="E8F1FB")
    section_font = Font(bold=True, color="0B2F4F")

    info = title_info or {}
    period_end = info.get("period_end")
    code6 = info.get("code6")
    provider = info.get("provider")
    stype = info.get("statement_type")
    pdf_url = info.get("pdf_url")
    pdf_path = info.get("pdf_path")

    def make_title(sheet_name: str) -> str:
        parts = []
        if code6:
            parts.append(str(code6))
        parts.append(sheet_name)
        if period_end:
            parts.append(f"报告期末日: {period_end}")
        if stype:
            parts.append(f"口径: {stype}")
        if provider:
            parts.append(f"数据源: {provider}")
        return " | ".join(parts)

    def make_note() -> str:
        # PDF 信息来源一致，放在注释行
        parts = []
        if pdf_url:
            parts.append(f"PDF链接: {pdf_url}")
        if pdf_path:
            parts.append(f"PDF本地路径: {pdf_path}")
        return " | ".join(parts) if parts else ""

    # 将 df 的缩进信息带到工作表里
    df_map = {
        "资产负债表": bs_df,
        "利润表": is_df,
        "现金流量表": cf_df,
    }

    for sname in ["资产负债表", "利润表", "现金流量表"]:
        ws = wb[sname]
        end_col = ws.max_column

        # 标题行 (row 1) + 注释行 (row 2)
        ws["A1"].value = make_title(sname)
        ws["A1"].fill = title_fill
        ws["A1"].font = title_font
        ws["A1"].alignment = title_alignment
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=end_col)
        ws.row_dimensions[1].height = 26

        note = make_note()
        if note:
            ws["A2"].value = note
            ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=end_col)
            ws["A2"].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            ws.row_dimensions[2].height = 34

        header_row = 3
        data_start = 4

        # 解析列位置（兼容列增减/顺序变化）
        headers = [ws.cell(header_row, c).value for c in range(1, end_col + 1)]

        def _col(name: str, default: int) -> int:
            try:
                return headers.index(name) + 1
            except Exception:
                return default

        subj_col = _col("科目", 1)
        value_col = _col("数值", end_col)

        # 冻结标题+注释+表头，并冻结 key+科目（即科目列之前的所有列）
        freeze_col = min(subj_col + 1, end_col)
        ws.freeze_panes = ws.cell(data_start, freeze_col).coordinate

        # 表头
        for cell in ws[header_row]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment

        # 缩进/分组样式
        df = df_map[sname]
        levels = df.get("__level")
        is_header = df.get("__is_header")

        for i in range(len(df)):
            excel_row = data_start + i
            subj_cell = ws.cell(excel_row, subj_col)
            val_cell = ws.cell(excel_row, value_col)

            lvl = int(levels.iloc[i]) if levels is not None else 0
            hdr = bool(is_header.iloc[i]) if is_header is not None else False

            if hdr:
                # 整行填充
                for c in range(1, end_col + 1):
                    ws.cell(excel_row, c).fill = section_fill

                subj_cell.font = section_font
                subj_cell.alignment = Alignment(horizontal="left", indent=0)
                # 标题行可能也有数值（如利润表“一、营业总收入”），不强行清空
                val_cell.alignment = Alignment(horizontal="right")
            else:
                subj_cell.alignment = Alignment(horizontal="left", indent=max(lvl, 0))

        # 交替底色（只对非标题行）
        for i in range(len(df)):
            excel_row = data_start + i
            hdr = bool(is_header.iloc[i]) if is_header is not None else False
            if hdr:
                continue
            if (i % 2) == 1:
                for c in range(1, end_col + 1):
                    ws.cell(excel_row, c).fill = zebra_fill

        # 数字列格式：对“数值”列做千分位
        for r in range(data_start, ws.max_row + 1):
            cell = ws.cell(r, value_col)
            if isinstance(cell.value, (int, float)):
                v = float(cell.value)
                if abs(v - round(v)) < 1e-9:
                    cell.number_format = "#,##0"
                else:
                    cell.number_format = "#,##0.00"
                cell.alignment = Alignment(horizontal="right")

        # 负数红色
        col_letter = ws.cell(header_row, value_col).column_letter
        rng = f"{col_letter}{data_start}:{col_letter}{ws.max_row}"
        ws.conditional_formatting.add(
            rng,
            CellIsRule(operator="lessThan", formula=["0"], font=Font(color="9C0006")),
        )

        # 列宽：自适应 + 关键列最小宽度兜底
        from openpyxl.utils import get_column_letter

        _autofit_worksheet(ws)

        v_letter = get_column_letter(value_col)
        s_letter = get_column_letter(subj_col)
        ws.column_dimensions[v_letter].width = max(ws.column_dimensions[v_letter].width or 0, 18)
        ws.column_dimensions[s_letter].width = max(ws.column_dimensions[s_letter].width or 0, 28)

        # key 列：避免过宽（key 很长时会把整张表拉得很宽）
        try:
            key_col = headers.index("key") + 1
            key_letter = get_column_letter(key_col)
            ws.column_dimensions[key_letter].width = min(ws.column_dimensions[key_letter].width or 0, 18)
        except Exception:
            pass

        # spacer 列宽（如果存在）
        try:
            spacer_col = headers.index(" ") + 1
            spacer_letter = get_column_letter(spacer_col)
            ws.column_dimensions[spacer_letter].width = 4
        except Exception:
            pass

    # meta sheet
    ws_meta = wb["META"] if "META" in wb.sheetnames else wb.create_sheet("META")
    ws_meta.delete_rows(1, ws_meta.max_row)
    ws_meta.append(["key", "value"])
    for k, v in meta.items():
        ws_meta.append([str(k), str(v)])
    ws_meta.freeze_panes = "A2"
    _autofit_worksheet(ws_meta)

    wb.save(out_path)

    return ExcelExportResult(path=str(out_path))

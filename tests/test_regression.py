from __future__ import annotations

from pathlib import Path

import pandas as pd

from finreport_fetcher.exporter.excel import export_bundle_to_excel
from finreport_fetcher.mappings.enrich import enrich_statement_df
from finreport_fetcher.providers.akshare_ths import AkshareThsProvider


def test_enrich_statement_df_dedup_lengths_and_keys():
    # Regression: previously we could drop some rows but still assign full-length keys/ens,
    # leading to "Length of values ... does not match length of index".
    df = pd.DataFrame(
        [
            {"科目": "负债合计", "数值": 10.0, "__level": 0, "__is_header": False},
            {"科目": "负债合计", "数值": 10.0, "__level": 0, "__is_header": False},  # redundant dup
            {"科目": "负债合计", "数值": 20.0, "__level": 0, "__is_header": False},  # meaningful dup
            {"科目": "固定资产", "数值": 50.0, "__level": 0, "__is_header": False},
            {"科目": "其中：固定资产", "数值": 0.0, "__level": 1, "__is_header": False},
            {"科目": "流动资产", "数值": None, "__level": 0, "__is_header": True},
            {"科目": "货币资金", "数值": 1.0, "__level": 1, "__is_header": False},
        ]
    )

    out = enrich_statement_df(df, sheet_name_cn="资产负债表")

    assert len(out) > 0
    assert "key" in out.columns
    assert "备注" in out.columns
    assert "英文" in out.columns
    assert "科目" in out.columns

    keys = out["key"].astype(str).tolist()
    # ASCII-only
    for k in keys:
        k.encode("ascii")

    # redundant duplicate should be removed
    assert keys.count("bs.total_liabilities") == 1
    # meaningful duplicate should remain with stable suffix
    assert any(k.startswith("bs.total_liabilities.dup") for k in keys)

    # tagged row should become stable sub key
    assert any(k == "bs.fixed_assets.sub" for k in keys)


def test_akshare_ths_balance_sheet_postprocess_inserts_headers():
    df = pd.DataFrame(
        [
            {"科目": "资产合计", "数值": 100.0, "__level": 0, "__is_header": False},
            {"科目": "负债合计", "数值": 40.0, "__level": 0, "__is_header": False},
            {"科目": "归属于母公司所有者权益合计", "数值": 60.0, "__level": 0, "__is_header": False},
            {"科目": "流动资产", "数值": None, "__level": 0, "__is_header": True},
            {"科目": "货币资金", "数值": 10.0, "__level": 1, "__is_header": False},
            {"科目": "非流动负债", "数值": None, "__level": 0, "__is_header": True},
            {"科目": "非流动负债合计", "数值": 5.0, "__level": 1, "__is_header": False},
            {"科目": "负债合计", "数值": 40.0, "__level": 1, "__is_header": False},
            {"科目": "实收资本（或股本）", "数值": 1.0, "__level": 1, "__is_header": False},
        ]
    )

    out = AkshareThsProvider._postprocess_balance_sheet(df)

    assert out.iloc[0]["科目"] == "报表核心指标"
    assert bool(out.iloc[0]["__is_header"]) is True

    # equity header should be inserted after the later total liabilities line
    names = out["科目"].astype(str).tolist()
    assert "股东权益" in names
    liab_i = max(i for i, n in enumerate(names) if n == "负债合计")
    eq_i = names.index("股东权益")
    assert eq_i == liab_i + 1


def test_export_excel_schema_and_autofit(tmp_path: Path):
    # Minimal bundle
    bs = pd.DataFrame([
        {"科目": "流动资产", "数值": None, "__level": 0, "__is_header": True},
        {"科目": "货币资金", "数值": 123.0, "__level": 1, "__is_header": False},
    ])
    inc = pd.DataFrame([
        {"科目": "营业总收入", "数值": 456.0, "__level": 0, "__is_header": False},
    ])
    cf = pd.DataFrame([
        {"科目": "经营活动产生的现金流量净额", "数值": 789.0, "__level": 0, "__is_header": False},
    ])

    out_path = tmp_path / "demo.xlsx"
    export_bundle_to_excel(
        out_path,
        balance_sheet=bs,
        income_statement=inc,
        cashflow_statement=cf,
        metrics_statement=None,
        meta={"k": "v"},
        title_info={
            "code6": "300454",
            "period_end": "2024-12-31",
            "statement_type": "merged",
            "provider": "akshare_ths",
            "pdf_url": "https://example.com/" + "x" * 200,  # intentionally long
        },
    )

    from openpyxl import load_workbook

    wb = load_workbook(out_path)
    ws = wb["资产负债表"]

    headers = [ws.cell(3, c).value for c in range(1, ws.max_column + 1)]
    assert headers[:6] == ["科目", "数值", " ", "key", "备注", "英文"]
    assert headers[-1] == "英文"

    # 科目列宽不应被 A1/A2 长标题撑爆
    subj_width = ws.column_dimensions["A"].width
    assert subj_width is not None
    # 不应被标题撑爆；同时需要足够宽避免遮挡
    assert 40.0 <= float(subj_width) <= 90.0

    wb.close()

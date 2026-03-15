from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

from finreport_fetcher.exporter.excel import export_bundle_to_excel


def test_excel_schema_is_fixed_and_subject_column_not_excessive(tmp_path: Path):
    # Minimal dfs with hierarchy columns
    bs = pd.DataFrame(
        {
            "科目": [
                "所有者权益合计",
                "资产总计",
                "负债合计",
                "归属于母公司所有者权益合计",
                "流动资产",
                "货币资金",
                "实收资本（或股本）",
            ],
            "数值": [100.0, 300.0, 200.0, 100.0, None, 123.0, 10.0],
            "__level": [0, 0, 0, 0, 0, 1, 1],
            "__is_header": [False, False, False, False, True, False, False],
        }
    )
    is_ = pd.DataFrame({"科目": ["营业总收入"], "数值": [1000.0], "__level": [0], "__is_header": [False]})
    cf = pd.DataFrame({"科目": ["经营活动产生的现金流量净额"], "数值": [1.0], "__level": [0], "__is_header": [False]})

    out_path = tmp_path / "t.xlsx"
    export_bundle_to_excel(
        out_path,
        balance_sheet=bs,
        income_statement=is_,
        cashflow_statement=cf,
        meta={"k": "v"},
        title_info={
            "code6": "000001",
            "period_end": "2024-12-31",
            "statement_type": "merged",
            "provider": "test",
            "pdf_url": "https://example.com/a.pdf",
            "pdf_path": "/tmp/a.pdf",
        },
    )

    wb = load_workbook(out_path)
    ws = wb["资产负债表"]

    # header row is row 3
    headers = [ws.cell(3, c).value for c in range(1, ws.max_column + 1)]
    assert headers == ["科目", "数值", " ", "key", "备注", "英文"]

    # subject column should not be blown up by title/note rows
    width_a = ws.column_dimensions["A"].width
    assert width_a is not None
    # 资产负债表：有最小宽度兜底，但不应被标题撑爆到极端
    assert 40.0 <= float(width_a) <= 90.0

    wb.close()

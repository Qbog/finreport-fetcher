from __future__ import annotations

from datetime import date
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

from finreport_fetcher.exporter.excel import export_bundle_to_excel
from finreport_fetcher.metrics_sheet import build_metrics_sheet


def test_metrics_sheet_keeps_all_section_rows_without_hash_suffix():
    source_df = pd.DataFrame([
        {"选项": "常用指标", "指标": "毛利率", "20241231": 91.2},
        {"选项": "盈利能力", "指标": "毛利率", "20241231": 91.2},
        {"选项": "常用指标", "指标": "总资产报酬率(ROA)", "20241231": 17.2},
        {"选项": "盈利能力", "指标": "总资产报酬率", "20241231": 17.1},
        {"选项": "盈利能力", "指标": "投入资本回报率", "20241231": 18.6},
    ])
    sheet = build_metrics_sheet(source_df, None, date(2024, 12, 31), "akshare")
    metric_rows = sheet[~sheet["key"].astype(str).str.startswith("metrics.section.")].copy()
    names = metric_rows["科目"].astype(str).tolist()
    keys = metric_rows["key"].astype(str).tolist()
    assert len(metric_rows) == 5
    assert names.count("毛利率") == 2
    assert "总资产报酬率(ROA)" in names
    assert "总资产报酬率" in names
    assert "投入资本回报率" in names
    assert "metrics.highlights.gross_margin" in keys
    assert "metrics.profitability.gross_margin" in keys
    assert "metrics.highlights.return_on_assets_roa" in keys
    assert "metrics.profitability.return_on_assets" in keys
    assert "metrics.profitability.return_on_invested_capital" in keys
    assert not any("__" in k for k in keys)


def test_build_metrics_sheet_and_export(tmp_path: Path):
    source_df = pd.DataFrame(
        [
            {
                "ts_code": "600519.SH",
                "ann_date": "20250328",
                "end_date": "20241231",
                "roe": 22.5,
                "roa": 17.2,
                "roic": 18.6,
                "ev": 1234.0,
                "ebitda": 234.5,
                "eps": 12.3,
                "grossprofit_margin": 89.0,
                "current_ratio": 3.2,
            }
        ]
    )
    metrics_df = source_df[["end_date", "ann_date", "roe", "roa", "roic", "ev", "ebitda"]].copy()
    sheet = build_metrics_sheet(source_df, metrics_df, date(2024, 12, 31), "tushare")
    assert not sheet.empty
    keys0 = sheet["key"].astype(str).tolist()
    assert "metrics.roe" in keys0
    assert "metrics.eps" in keys0
    assert "metrics.current_ratio" in keys0
    assert not any("__" in k for k in keys0)
    english_values = [str(x) for x in sheet["英文"].dropna().tolist() if str(x).strip()]
    assert "基本每股收益(EPS)" not in english_values
    assert any("Basic Eps" == x or "Return on equity (ROE)" == x for x in english_values)
    assert sheet["科目"].astype(str).tolist().count("净资产收益率(ROE)") == 1

    bs = pd.DataFrame({"科目": ["资产总计"], "数值": [100.0], "__level": [0], "__is_header": [False]})
    is_ = pd.DataFrame({"科目": ["营业总收入"], "数值": [50.0], "__level": [0], "__is_header": [False]})
    cf = pd.DataFrame({"科目": ["经营活动产生的现金流量净额"], "数值": [20.0], "__level": [0], "__is_header": [False]})
    out = tmp_path / "demo.xlsx"
    export_bundle_to_excel(
        out,
        balance_sheet=bs,
        income_statement=is_,
        cashflow_statement=cf,
        metrics_statement=sheet,
        meta={},
        title_info={"code6": "600519", "period_end": "2024-12-31", "provider": "akshare", "metrics_provider": "akshare"},
    )
    wb = load_workbook(out)
    assert "财报指标" in wb.sheetnames
    ws = wb["财报指标"]
    headers = [ws.cell(3, c).value for c in range(1, ws.max_column + 1)]
    assert headers == ["科目", "数值", " ", "key", "备注", "英文"]
    keys = [ws.cell(r, 4).value for r in range(4, ws.max_row + 1) if ws.cell(r, 4).value]
    assert "metrics.roe" in keys
    wb.close()

from __future__ import annotations

from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

from finreport_fetcher.exporter.excel import export_bundle_to_excel
from finreport_fetcher.metrics_sheet import build_placeholder_metrics_sheet


def test_export_always_keeps_metrics_sheet_when_placeholder_provided(tmp_path: Path):
    bs = pd.DataFrame({"科目": ["资产总计"], "数值": [100.0], "__level": [0], "__is_header": [False]})
    is_ = pd.DataFrame({"科目": ["营业总收入"], "数值": [50.0], "__level": [0], "__is_header": [False]})
    cf = pd.DataFrame({"科目": ["经营活动产生的现金流量净额"], "数值": [20.0], "__level": [0], "__is_header": [False]})
    out = tmp_path / "demo.xlsx"
    export_bundle_to_excel(
        out,
        balance_sheet=bs,
        income_statement=is_,
        cashflow_statement=cf,
        metrics_statement=build_placeholder_metrics_sheet("测试占位"),
        meta={},
        title_info={"code6": "600519", "period_end": "2024-12-31", "provider": "akshare"},
    )
    wb = load_workbook(out, read_only=True)
    assert "财报指标" in wb.sheetnames
    wb.close()
